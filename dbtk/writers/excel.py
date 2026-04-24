# dbtk/writers/excel.py
"""
Excel writer for database results using openpyxl.
"""
import logging
from typing import Any, Union, List, Optional, Iterable, Dict, TYPE_CHECKING
from pathlib import Path
from datetime import datetime, date, time
from zipfile import BadZipFile
import fnmatch
import hashlib

from .base import BatchWriter, RecordLike

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, NamedStyle
    from openpyxl.comments import Comment
    from openpyxl.utils.exceptions import InvalidFileException
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    InvalidFileException = Exception  # fallback so except clause is valid

logger = logging.getLogger(__name__)

MIDNIGHT = time(0, 0)

_BUILTIN_STYLE_NAMES = frozenset({
    'date_style', 'datetime_style', 'hyperlink_style',
    'bold_style', 'header_vert_style',
    'currency_style', 'percent_style', 'comma_style',
})


class ExcelWriter(BatchWriter):
    """
    Stateful Excel writer using openpyxl.

    Keeps the workbook open across multiple write_batch() calls and saves only on context exit.
    Designed for both single-sheet legacy use and multi-sheet reports.

    Supports all 3 BatchWriter modes:
    1. Complete write from __init__ + write()
    2. Batch write (no data on init) + write_batch()
    3. Hybrid: data on init + write() + write_batch()

    Usage examples::

        # Mode 1: Traditional single-shot write
        ExcelWriter(cursor, 'report.xlsx').write()

        # Mode 2: Pure streaming with write_batch()
        with ExcelWriter(file='report.xlsx') as writer:
            writer.write_batch(cursor)  # goes to sheet 'Data'

        # Mode 3: Hybrid - initial data + streaming
        with ExcelWriter(first_batch, 'report.xlsx') as writer:
            writer.write()  # Write initial batch
            writer.write_batch(second_batch)  # Stream additional batches

        # Multi-sheet report
        with ExcelWriter(file='report.xlsx', sheet_name='Summary') as writer:
            writer.write_batch(summary_data, sheet_name='Summary')
            writer.write_batch(users_data, sheet_name='Users')
            writer.write_batch(orders_data, sheet_name='Orders')

        # Streaming / batch mode
        with ExcelWriter(file='large.xlsx') as writer:
            for batch in large_generator:
                writer.write_batch(batch, sheet_name='Data')  # appends to 'Data'
    """

    accepts_file_handle = False
    preserve_types = True

    def __init__(
        self,
        data: Optional[Iterable[RecordLike]] = None,
        file: Optional[Union[str, Path]] = None,
        sheet_name: Optional[str] = None,
        headers: Optional[List[str]] = None,
        write_headers: bool = True,
        formatting: Optional[Dict] = None,
    ):
        """
        Initialize the Excel writer.

        Parameters
        ----------
        data : Iterable[RecordLike], optional
            Initial data to write. If None, use write_batch() for streaming mode.
        file : str or Path, optional
            Output Excel file (.xlsx). Required for Excel output.
        sheet_name : str, optional
            Default/active sheet name to use for write_batch() calls without explicit sheet_name
        headers : List[str], optional
            Header row text. If None, checks data.description for original column names,
            then falls back to detected column names. Useful when Record field names have been
            normalized (e.g., lowercased) but you want original database column names in Excel.
        write_headers : bool, default True
            Whether to write column headers (only when sheet is empty)
        formatting : dict, optional
            Worksheet formatting rules. Supported keys:

            * ``'styles'`` — named style definitions, e.g.
              ``{'fmt_fees': {'bg_color': '#d5f1cc'}}``
            * ``'columns'`` — wildcard pattern → properties, e.g.
              ``{'fees*': {'format': 'fmt_fees', 'width': 20}, 'resv_*': {'hidden': 1}}``
              Later patterns override earlier ones per property. Matching is case-insensitive.
              Properties: ``format`` (style name or inline dict applied to data cells),
              ``header_format`` (style name or inline dict applied to the header cell only;
              owns the cell entirely so include ``font: {bold: True}`` if needed),
              ``width`` (float), ``hidden`` (0/1), ``comment`` (string — adds an Excel
              comment/note to the header cell).
              Built-in style names available without defining in ``styles``:
              ``bold_style``, ``header_vert_style``, ``date_style``, ``datetime_style``,
              ``hyperlink_style``, ``currency_style``, ``percent_style``, ``comma_style``.
            * ``'rows'`` — row index → properties dict. Index 0 = header row. Positive
              integers are 1-based data row indices. ``'style'`` key accepts a callable
              ``lambda rec: style_name_or_None`` applied to every data row. ``'odd'``
              and ``'even'`` keys accept ``{'format': style_name}`` dicts applied
              automatically to odd (1st, 3rd, …) and even (2nd, 4th, …) data rows.
            * ``'min_column_width'`` — minimum column width in Excel units applied to
              all auto-sized columns (default ``6``). Lower this for narrow indicator
              columns (``'Y'``/``'N'``, flags) where ``3``–``4`` is sufficient.
              Explicit ``width`` values in column rules are not affected.
            * ``'max_column_width'`` — maximum column width in Excel units applied to
              all auto-sized columns (default ``60``). Explicit ``width`` values in
              column rules are not affected.
            * ``'auto_filter'`` — if truthy, enables Excel's dropdown auto-filter on
              the header row.
            * ``'freeze'`` — cell reference string for freeze panes, e.g. ``'D2'``.
              Defaults to ``'A2'``. Pass ``None`` to disable freezing.
            * ``'header_auto_rotate'`` — automatically apply ``header_vert_style`` to
              columns whose header text is significantly longer than their sampled data.
              Pass a float ratio (e.g. ``1.5``) or a dict
              ``{'ratio': 1.5, 'min_length': 8, 'height_factor': 6.5}``. Both conditions must hold: header
              length ≥ ``min_length`` (default 8) **and** header length > data width ×
              ``ratio`` (default 1.5). Header row height is computed automatically from
              the longest rotated header (≈ 6.5 pt/char) unless ``rows[0]['height']``
              is set explicitly. Columns with an explicit ``header_format`` are excluded
              from auto-rotation.

            Style property dicts support: ``bg_color`` (hex string), ``font`` (dict of
            Font kwargs), ``number_format`` (string), ``alignment`` (dict of Alignment
            kwargs).
        """
        if not HAS_OPENPYXL:
            raise ImportError("ExcelWriter requires openpyxl: pip install openpyxl")
        if file is None:
            raise ValueError("ExcelWriter requires an output file path")

        super().__init__(data=data, file=file, headers=headers, write_headers=write_headers)

        self.output_path = Path(file)
        self.active_sheet: Optional[str] = sheet_name
        self.workbook: Optional[Workbook] = None
        self._sheets_written_this_session: set = set()  # Track sheets written in this session
        self.formatting = formatting or {}

        self._load_or_create_workbook()

    def _load_or_create_workbook(self) -> None:
        """Load existing workbook or create a new one."""
        try:
            if self.output_path.exists():
                with open(self.output_path, mode='r+b'):
                    # make sure it is actually writable (not open in Excel)
                    pass
                self.workbook = load_workbook(self.output_path)
                logger.info(f"Loaded existing workbook: {self.output_path}")
            else:
                with open(self.output_path, mode='wb'):
                    pass
                self.workbook = Workbook()
                if 'Sheet' in self.workbook.sheetnames:
                    self.workbook.remove(self.workbook['Sheet'])
        except (InvalidFileException, BadZipFile, ValueError) as e:
            raise ValueError(
                f"File '{self.output_path}' exists but is not a valid Excel workbook. "
                f"Original error: {e}"
            ) from e
        except PermissionError:
            raise PermissionError(
                f"Cannot write to '{self.output_path}' - file may be open in Excel or another application. "
                "Please close the file and try again."
            )

        self._register_styles()

    def _register_styles(self) -> None:
        """Register built-in styles available to all sheets by name.

        Built-in styles
        ---------------
        date_style
            Date number format: ``YYYY-MM-DD``.
        datetime_style
            Datetime number format: ``YYYY-MM-DD HH:MM:SS``.
        hyperlink_style
            Blue underlined font (used automatically by LinkedExcelWriter).
        bold_style
            Bold font. Useful as a ``header_format`` when you only want emphasis.
        header_vert_style
            Bold font + 90° text rotation. Pair with ``rows: {0: {'height': 120}}``
            for narrow rotated column headers.
        currency_style
            Number format: ``#,##0.00``.
        percent_style
            Number format: ``0.00%``.
        comma_style
            Number format: ``#,##0``.
        """
        if self.workbook is None:
            return

        from openpyxl.styles import Alignment as _Alignment
        _bold_font = Font(bold=True)
        styles = [
            NamedStyle(name='date_style', number_format='YYYY-MM-DD'),
            NamedStyle(name='datetime_style', number_format='YYYY-MM-DD HH:MM:SS'),
            NamedStyle(
                name='hyperlink_style',
                font=Font(color="0000FF", underline="single")
            ),
            NamedStyle(name='bold_style', font=_bold_font),
            NamedStyle(name='header_vert_style', font=_bold_font,
                       alignment=_Alignment(text_rotation=90, horizontal='center')),
            NamedStyle(name='currency_style', number_format='#,##0.00'),
            NamedStyle(name='percent_style', number_format='0.00%'),
            NamedStyle(name='comma_style', number_format='#,##0'),
        ]

        for style in styles:
            if style.name not in self.workbook.named_styles:
                self.workbook.add_named_style(style)

        for style_name, props in self.formatting.get('styles', {}).items():
            if style_name in _BUILTIN_STYLE_NAMES:
                logger.warning(
                    f"Style name '{style_name}' is reserved by ExcelWriter and cannot be "
                    "overridden; rename your style to apply it."
                )
                continue
            if style_name not in self.workbook.named_styles:
                self.workbook.add_named_style(self._build_named_style(style_name, props))

    @staticmethod
    def _build_named_style(name: str, props: dict) -> 'NamedStyle':
        """Build a NamedStyle from a properties dict."""
        from openpyxl.styles import PatternFill, Alignment
        style = NamedStyle(name=name)
        if 'bg_color' in props:
            color = props['bg_color'].lstrip('#')
            style.fill = PatternFill(fill_type='solid', fgColor=color)
        if 'font' in props:
            style.font = Font(**props['font'])
        if 'number_format' in props:
            style.number_format = props['number_format']
        if 'alignment' in props:
            style.alignment = Alignment(**props['alignment'])
        return style

    def _ensure_style(self, props: dict) -> str:
        """Register an inline format dict as a NamedStyle; return its name."""
        key = tuple(sorted((k, str(v)) for k, v in props.items()))
        name = 'fmt_' + hashlib.md5(str(key).encode()).hexdigest()[:8]
        if name not in self.workbook.named_styles:
            self.workbook.add_named_style(self._build_named_style(name, props))
        return name

    def _build_col_fmt_map(self, columns: List[str]) -> list:
        """Build per-column formatting list from wildcard pattern rules.

        Returns a list (indexed by 0-based column position) of property dicts.
        Patterns are applied in definition order; later rules win per property.
        Matching is case-insensitive fnmatch glob.
        """
        col_rules = self.formatting.get('columns', {})
        if not col_rules:
            return []
        result: List[dict] = [{} for _ in columns]
        for pattern, props in col_rules.items():
            pattern_lower = pattern.lower()
            for col_idx, col_name in enumerate(columns):
                if fnmatch.fnmatch(col_name.lower(), pattern_lower):
                    result[col_idx].update(props)
        for col_props in result:
            fmt = col_props.get('format')
            if isinstance(fmt, dict):
                col_props['format'] = self._ensure_style(fmt)
            hfmt = col_props.get('header_format')
            if isinstance(hfmt, dict):
                col_props['header_format'] = self._ensure_style(hfmt)
        return result

    def _finalize_headers(
        self,
        worksheet: 'Worksheet',
        header_widths: list,
        data_widths: list,
        col_fmt: list,
        rows_fmt: dict,
        link_mapping: Optional[dict] = None,
    ) -> None:
        """Apply column widths, auto-rotate, header height, and freeze panes.

        Called once after data has been sampled, so auto-rotate decisions have
        accurate data-width information. ``link_mapping`` is LinkedExcelWriter's
        column → (LinkSource, mode) map; LinkSource display widths substitute for
        sampled data widths on linked columns.
        """
        link_mapping = link_mapping or {}

        # Substitute LinkSource display widths for linked columns
        effective_data_widths = list(data_widths)
        for col_idx, col_name in enumerate(self.columns, 1):
            if col_name in link_mapping:
                source, _ = link_mapping[col_name]
                if source.max_display_width > 0:
                    effective_data_widths[col_idx - 1] = source.max_display_width

        # Auto-rotate: detect columns whose header is significantly longer than their data
        auto_rotated: set = set()
        har_height_factor = 6.5
        har = self.formatting.get('header_auto_rotate')
        if har:
            if isinstance(har, dict):
                har_min = har.get('min_length', 8)
                har_ratio = har.get('ratio', 1.5)
                har_height_factor = har.get('height_factor', 6.5)
            else:
                har_min = 8
                har_ratio = float(har)
                har_height_factor = 6.5

            for col_idx, (hw, dw) in enumerate(zip(header_widths, effective_data_widths), 1):
                col_props = col_fmt[col_idx - 1] if col_fmt else {}
                if col_props.get('header_format'):
                    continue  # explicit header_format takes precedence
                if hw >= har_min and hw > dw * har_ratio:
                    auto_rotated.add(col_idx)
                    worksheet.cell(1, col_idx).style = 'header_vert_style'

        # Column widths: auto-rotated columns use data width only; others use max of both
        min_col_width = self.formatting.get('min_column_width', 6)
        max_col_width = self.formatting.get('max_column_width', 60)
        for col_idx, (hw, dw) in enumerate(zip(header_widths, effective_data_widths), 1):
            raw = dw if col_idx in auto_rotated else max(hw, dw)
            adjusted = min(max(raw + 2, min_col_width), max_col_width)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted

        # User column-rule overrides (width, hidden, comment)
        for col_idx, col_props in enumerate(col_fmt, 1):
            if col_props:
                col_letter = get_column_letter(col_idx)
                if 'width' in col_props:
                    worksheet.column_dimensions[col_letter].width = col_props['width']
                if 'hidden' in col_props:
                    worksheet.column_dimensions[col_letter].hidden = bool(col_props['hidden'])
                if 'comment' in col_props:
                    worksheet.cell(1, col_idx).comment = Comment(col_props['comment'], '')

        # Header row height: explicit rows[0] wins; otherwise auto from rotated header lengths
        explicit_h = None
        if isinstance(rows_fmt, dict) and 0 in rows_fmt:
            rp = rows_fmt[0]
            explicit_h = rp.get('height') if isinstance(rp, dict) else None
        if explicit_h is not None:
            worksheet.row_dimensions[1].height = explicit_h
        elif auto_rotated:
            max_rotated_len = max(header_widths[i - 1] for i in auto_rotated)
            worksheet.row_dimensions[1].height = max_rotated_len * har_height_factor

        # Freeze panes
        freeze = self.formatting.get('freeze', 'A2')
        if freeze:
            worksheet.freeze_panes = freeze

        # Auto-filter on header row
        if self.formatting.get('auto_filter'):
            last_col = get_column_letter(len(self.columns))
            worksheet.auto_filter.ref = f"A1:{last_col}1"

    def _get_or_create_worksheet(self, sheet_name: str) -> 'Worksheet':
        """Get existing worksheet or create new one."""
        from openpyxl.worksheet.worksheet import Worksheet

        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            return self.workbook.create_sheet(sheet_name)

    def _clear_worksheet(self, worksheet: 'Worksheet') -> None:
        """Clear all rows from a worksheet."""
        # Simply delete all rows - openpyxl will handle this correctly
        if worksheet.max_row >= 1:
            worksheet.delete_rows(1, worksheet.max_row)

    def _get_named_style(self, name: str) -> 'NamedStyle':
        for style in self.workbook.named_styles:
            if style.name == name:
                return style
        raise KeyError(f"Named style '{name}' not found")

    def _write_to_worksheet(
        self,
        data: Iterable[RecordLike],
        worksheet: 'Worksheet',
        columns: Optional[List[str]] = None,
        write_headers: bool = True,

    ) -> int:
        """
        Internal method: write data to an already-selected worksheet.

        Returns number of rows written.
        """
        from openpyxl.worksheet.worksheet import Worksheet

        # Lazy init columns
        self.data_iterator, detected_columns = self._get_data_iterator(data, columns)
        self.columns = detected_columns

        if not self.columns:
            raise ValueError("Could not determine columns from data")

        col_fmt = self._build_col_fmt_map(self.columns)
        rows_fmt = self.formatting.get('rows', {})
        row_style_fn = rows_fmt.get('style') if isinstance(rows_fmt, dict) else None
        odd_style = rows_fmt.get('odd', {}).get('format') if isinstance(rows_fmt, dict) else None
        even_style = rows_fmt.get('even', {}).get('format') if isinstance(rows_fmt, dict) else None

        row_count = 0
        header_widths = [len(col) for col in self.columns]
        data_widths = [0] * len(self.columns)
        width_sample_size = 15
        header_font = Font(bold=True)

        should_write_headers = write_headers and worksheet.cell(1, 1).value is None
        data_start_row = 2 if should_write_headers else worksheet.max_row + 1

        if should_write_headers:
            for col_idx, column_name in enumerate(self._get_headers(data), 1):
                cell = worksheet.cell(row=1, column=col_idx, value=column_name)
                hfmt = col_fmt[col_idx - 1].get('header_format') if col_fmt else None
                if hfmt:
                    cell.style = hfmt
                else:
                    cell.font = header_font

        # Write data rows
        for row_idx, record in enumerate(self.data_iterator, data_start_row):
            values = self._row_to_tuple(record)
            data_row_num = row_idx - data_start_row + 1
            alt_style = odd_style if data_row_num % 2 else even_style
            row_style = row_style_fn(record) if callable(row_style_fn) else alt_style

            if isinstance(rows_fmt, dict) and data_row_num in rows_fmt:
                rp = rows_fmt[data_row_num]
                if isinstance(rp, dict) and 'height' in rp:
                    worksheet.row_dimensions[row_idx].height = rp['height']

            for col_idx, value in enumerate(values, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                is_date_val = isinstance(value, (date, datetime))

                if isinstance(value, datetime) and value.time() != MIDNIGHT:
                    cell.value = value
                    cell.style = 'datetime_style'
                    if row_count < width_sample_size:
                        data_widths[col_idx - 1] = max(data_widths[col_idx - 1], 19)
                elif isinstance(value, (date, datetime)):
                    cell.value = value
                    cell.style = 'date_style'
                    if row_count < width_sample_size:
                        data_widths[col_idx - 1] = max(data_widths[col_idx - 1], 10)
                elif value is None:
                    cell.value = ''
                else:
                    cell.value = value
                    if row_count < width_sample_size:
                        data_widths[col_idx - 1] = max(data_widths[col_idx - 1], len(str(value)))

                if not is_date_val and col_fmt:
                    col_style = col_fmt[col_idx - 1].get('format')
                    if col_style:
                        cell.style = col_style

                if row_style and not is_date_val:
                    cell.style = row_style

            row_count += 1

        if should_write_headers:
            self._finalize_headers(worksheet, header_widths, data_widths, col_fmt, rows_fmt)

        return row_count

    def write_batch(
        self,
        data: Iterable[RecordLike],
        sheet_name: Optional[str] = None,
    ) -> None:
        """
        Write a batch of data to a sheet.

        If this is the first write to this sheet in the current session, the sheet
        is cleared first. Subsequent writes to the same sheet append data.

        Parameters
        ----------
        data : Iterable[RecordLike]
            The data batch
        sheet_name : str, optional
            Target sheet. If None, uses active_sheet or defaults to 'Data'
        """
        if self.workbook is None:
            raise RuntimeError("Workbook not initialized")

        target_sheet = sheet_name or self.active_sheet or 'Data'
        if sheet_name:
            self.active_sheet = sheet_name

        worksheet = self._get_or_create_worksheet(target_sheet)

        # Clear sheet if this is the first write to it in this session
        if target_sheet not in self._sheets_written_this_session:
            self._clear_worksheet(worksheet)
            self._sheets_written_this_session.add(target_sheet)

        row_count = self._write_to_worksheet(
            data=data,
            worksheet=worksheet,
            write_headers=self.write_headers
        )

        self._row_num += row_count
        logger.info(f"Wrote {row_count} rows to sheet '{target_sheet}' (total: {self._row_num})")

    def _write_data(self, file_obj: Any) -> None:
        """
        BatchWriter contract implementation.

        Writes current data_iterator (set up by _lazy_init) to the active sheet or 'Data'.
        This is called by write() when data was provided at initialization.
        """
        if self.data_iterator is None:
            raise RuntimeError("No data provided")

        if not self.columns:
            raise RuntimeError("Columns not initialized")

        if self.workbook is None:
            raise RuntimeError("Workbook not initialized")

        col_fmt = self._build_col_fmt_map(self.columns)
        rows_fmt = self.formatting.get('rows', {})
        row_style_fn = rows_fmt.get('style') if isinstance(rows_fmt, dict) else None
        odd_style = rows_fmt.get('odd', {}).get('format') if isinstance(rows_fmt, dict) else None
        even_style = rows_fmt.get('even', {}).get('format') if isinstance(rows_fmt, dict) else None

        target_sheet = self.active_sheet or 'Data'
        worksheet = self._get_or_create_worksheet(target_sheet)

        should_write_headers = self.write_headers and not self._headers_written and worksheet.cell(1, 1).value is None
        data_start_row = 2 if should_write_headers else worksheet.max_row + 1

        if should_write_headers:
            header_font = Font(bold=True)
            for col_idx, column_name in enumerate(self._get_headers(), 1):
                cell = worksheet.cell(row=1, column=col_idx, value=column_name)
                hfmt = col_fmt[col_idx - 1].get('header_format') if col_fmt else None
                if hfmt:
                    cell.style = hfmt
                else:
                    cell.font = header_font
            self._headers_written = True

        row_count = 0
        header_widths = [len(col) for col in self.columns]
        data_widths = [0] * len(self.columns)
        width_sample_size = 15

        # Write data rows
        for row_idx, record in enumerate(self.data_iterator, data_start_row):
            values = self._row_to_tuple(record)
            data_row_num = row_idx - data_start_row + 1
            alt_style = odd_style if data_row_num % 2 else even_style
            row_style = row_style_fn(record) if callable(row_style_fn) else alt_style

            if isinstance(rows_fmt, dict) and data_row_num in rows_fmt:
                rp = rows_fmt[data_row_num]
                if isinstance(rp, dict) and 'height' in rp:
                    worksheet.row_dimensions[row_idx].height = rp['height']

            for col_idx, value in enumerate(values, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                is_date_val = isinstance(value, (date, datetime))

                if isinstance(value, datetime) and value.time() != MIDNIGHT:
                    cell.value = value
                    cell.style = 'datetime_style'
                    if row_count < width_sample_size:
                        data_widths[col_idx - 1] = max(data_widths[col_idx - 1], 19)
                elif isinstance(value, (date, datetime)):
                    cell.value = value
                    cell.style = 'date_style'
                    if row_count < width_sample_size:
                        data_widths[col_idx - 1] = max(data_widths[col_idx - 1], 10)
                elif value is None:
                    cell.value = ''
                else:
                    cell.value = value
                    if row_count < width_sample_size:
                        data_widths[col_idx - 1] = max(data_widths[col_idx - 1], len(str(value)))

                if not is_date_val and col_fmt:
                    col_style = col_fmt[col_idx - 1].get('format')
                    if col_style:
                        cell.style = col_style

                if row_style and not is_date_val:
                    cell.style = row_style

            row_count += 1

        if should_write_headers:
            self._finalize_headers(worksheet, header_widths, data_widths, col_fmt, rows_fmt)

        self._row_num += row_count
        logger.info(f"Wrote {row_count} rows to sheet '{target_sheet}' (total: {self._row_num})")

    def _save_workbook(self):
        """Save the workbook. Idempotent - safe to call multiple times."""
        if self.workbook is not None:
            try:
                self.workbook.save(self.output_path)
                logger.info(f"Saved workbook: {self.output_path}")
            except Exception as e:
                logger.error(f"Failed to save workbook: {e}")
                raise
            finally:
                self.workbook = None  # Mark as saved to prevent duplicate saves

    def close(self):
        """Close the writer and save the workbook."""
        self._save_workbook()

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Save workbook on context exit."""
        self.close()
        return False


def to_excel(
    data,
    file: Union[str, Path],
    sheet: str = 'Data',
    headers: Optional[List[str]] = None,
    write_headers: bool = True,
) -> None:
    """
    Legacy convenience function — writes a single sheet.

    Parameters
    ----------
    data : Iterable[RecordLike]
        Data to write (cursor, list of Records, etc.)
    file : str or Path
        Output Excel file (.xlsx)
    sheet : str, default 'Data'
        Sheet name to write to
    headers : List[str], optional
        Header row text. If None, uses cursor.description or detected column names
    write_headers : bool, default True
        Whether to write column headers

    Examples
    --------
    # Write cursor with original database column names
    to_excel(cursor, 'report.xlsx')

    # Override header names
    to_excel(cursor, 'report.xlsx', headers=['User ID', 'Full Name', 'Email'])

    For multi-sheet or advanced reports, use ExcelWriter as a context manager with write_batch().
    """
    with ExcelWriter(data=None, file=file, headers=headers, write_headers=write_headers) as writer:
        writer.write_batch(data=data, sheet_name=sheet)


class LinkSource:
    """
    Defines a linkable data source for creating internal and external hyperlinks in Excel.

    LinkSource enables rich hyperlinking between worksheets and to external URLs. Each
    LinkSource is registered once with LinkedExcelWriter and can be used across multiple
    sheets to create consistent, navigable Excel reports.

    As the source sheet is written, LinkSource caches each record's cell reference and
    generates formatted link text and URLs using Python's str.format_map(). Later sheets
    can reference these cached records to create hyperlinks.

    Link Types
    ----------
    * **Internal links** - Excel cell references like ``#Students!A5`` for navigation
    * **External links** - HTTP/HTTPS URLs like ``https://crm.example.com/contact/123``
    * **Hybrid mode** - Stores both internal and external, defaults to external if available

    Parameters
    ----------
    name : str
        Unique identifier for this link source. Used when registering links in
        write_batch() (e.g., ``links={"student_name": "student"}``).
    source_sheet : str, optional
        The worksheet name that serves as the authoritative source for this entity.
        Internal links will point to rows in this sheet. This sheet must be written
        before any sheets that reference it. Not required when external_only=True.
    key_column : str, optional
        The column name containing unique identifiers for records (e.g., "student_id").
        This column must exist in both the source sheet data and any sheets that
        create links to it. Not required when external_only=True.
    url_template : str, optional
        Python format string for generating external URLs. Uses str.format_map() with
        the full record dict as context. Example: ``"https://app.com/users/{user_id}"``
    text_template : str, optional
        Python format string for generating link display text. Uses str.format_map()
        with the full record dict. Example: ``"{last_name}, {first_name} ({dept})"``
        If not provided, uses the column value.
    missing_text : str, optional
        Fallback text to display when a link target cannot be resolved. If None,
        displays the raw value from the detail row.
    external_only : bool, default False
        If True, this LinkSource generates external links directly from current row data
        without caching. Can be reused across multiple sheets. source_sheet and key_column
        are not required. If False (default), caches records for cross-sheet linking.

    Attributes
    ----------
    _records : dict
        Internal cache mapping key values to link metadata (ref, display_text, url).
        Populated automatically as the source sheet is written.

    Examples
    --------
    **Internal links only (sheet navigation)**::

        student_link = LinkSource(
            name="student",
            source_sheet="Students",
            key_column="student_id"
        )

    **External links with custom text**::

        employee_link = LinkSource(
            name="employee",
            source_sheet="Employees",
            key_column="employee_id",
            url_template="https://hr.company.com/profile/{employee_id}",
            text_template="{last_name}, {first_name} ({department})"
        )

    **Hybrid with missing value handling**::

        customer_link = LinkSource(
            name="customer",
            source_sheet="Customers",
            key_column="customer_id",
            url_template="https://crm.company.com/customers/{crm_id}",
            text_template="{company_name} - {contact_name}",
            missing_text="[Unknown Customer]"
        )

    **External-only links (reusable across sheets)**::

        # For external links - reusable on any sheet with required columns
        imdb_link = LinkSource(
            name="imdb",
            url_template="https://imdb.com/title/{tconst}",
            text_template="{primary_title} ({start_year})",
            external_only=True  # No source_sheet needed - works on any sheet
        )
        # Use on multiple sheets with same columns:
        writer.write_batch(movies, "Movies", links={"primary_title": "imdb"})
        writer.write_batch(top_rated, "Top Rated", links={"primary_title": "imdb"})

    Notes
    -----
    * The source sheet MUST be written before sheets that reference it
    * All template fields must exist in the record data or KeyError will be logged
    * Key values are converted to strings for cache lookups
    * Templates use Python's str.format_map() - use double braces {{}} to escape

    See Also
    --------
    LinkedExcelWriter : Writer that uses LinkSource for hyperlinking
    LinkedExcelWriter.register_link_source : Method to register a LinkSource
    """
    def __init__(self,
                 name: str,
                 source_sheet: str = None,
                 key_column: str = None,
                 url_template: str = None,
                 text_template: str = None,
                 missing_text: str = None,
                 external_only: bool = False):
        self.name = name
        self.source_sheet = source_sheet
        self.key_column = key_column
        self.url_template = url_template
        self.text_template = text_template
        self.missing_text = missing_text
        self.external_only = external_only
        self._records = {}

        # Validation
        if external_only:
            if not url_template:
                raise ValueError(f"url_template is required when external_only=True for LinkSource '{name}'")
        else:
            if not source_sheet:
                raise ValueError(f"source_sheet is required when external_only=False for LinkSource '{name}'")
            if not key_column:
                raise ValueError(f"key_column is required when external_only=False for LinkSource '{name}'")

        # Track display width for column sizing (sample first 100 rows, cap at 50 chars)
        self._max_display_width: int = 0
        self._sample_count: int = 0

    def cache_record(self, key_value: Any, row_dict: Dict[str, Any], ref: str) -> None:
        """Cache a record for cross-sheet linking (unless external_only=True)."""
        # Skip caching if this LinkSource is external-only (self-linking only)
        if self.external_only:
            # Still track display width for column sizing
            if self._sample_count < 100:
                if self.text_template:
                    try:
                        display_text = self.text_template.format_map(row_dict)
                    except KeyError:
                        display_text = str(key_value)
                else:
                    display_text = str(key_value)

                display_len = min(len(display_text), 50)
                self._max_display_width = max(self._max_display_width, display_len)
                self._sample_count += 1
            return

        key_str = str(key_value)

        if self.text_template:
            try:
                display_text = self.text_template.format_map(row_dict)
            except KeyError as e:
                logger.warning(f"Missing key {e} in text_template for {self.name}")
                display_text = f"{key_value} ({self.missing_text})"
        else:
            # No template → use the raw key value (clean, expected)
            display_text = str(key_value)

        # Track max display width for first 100 records (capped at 50 chars)
        if self._sample_count < 100:
            display_len = min(len(display_text), 50)
            self._max_display_width = max(self._max_display_width, display_len)
            self._sample_count += 1

        record = {
            "ref": ref,
            "display_text": display_text,
        }

        if self.url_template:
            try:
                record["url"] = self.url_template.format_map(row_dict)
            except KeyError as e:
                logger.warning(f"Missing key {e} in url_template for {self.name}")

        if key_str not in self._records:
            self._records[key_str] = record

    def generate_link_from_row(
        self,
        row_dict: Dict[str, Any],
        ref: str,
        mode: str = "external",
        column_value: Any = None
    ) -> Optional[dict]:
        """
        Generate link info directly from row data (for self-linking or external-only).

        Used when writing the source sheet itself to create links from current row
        instead of looking up from cache.

        Parameters
        ----------
        row_dict : dict
            The current row's data as a dictionary
        ref : str
            The cell reference for internal links (e.g., "#Movies!A5")
        mode : str
            "external" or "internal"
        column_value : Any, optional
            For external_only sources, the value from the linked column

        Returns
        -------
        dict or None
            Dict with "target" and "display_text", or None if link cannot be generated
        """
        # Generate display text
        if self.text_template:
            try:
                display_text = self.text_template.format_map(row_dict)
            except KeyError as e:
                logger.warning(f"Missing key {e} in text_template for {self.name}")
                return None
        else:
            # No template → use column_value for external_only, or key_column for others
            if self.external_only:
                # For external_only: use the linked column's value
                if column_value is None:
                    return None
                display_text = str(column_value)
            elif self.key_column:
                # For self-linking: use key_column
                key_value = row_dict.get(self.key_column)
                if key_value is None:
                    return None
                display_text = str(key_value)
            else:
                return None

        # Track max display width for first 100 self-links (capped at 50 chars)
        if self._sample_count < 100:
            display_len = min(len(display_text), 50)
            self._max_display_width = max(self._max_display_width, display_len)
            self._sample_count += 1

        # Generate target
        if mode == "external":
            if self.url_template:
                try:
                    target = self.url_template.format_map(row_dict)
                except KeyError as e:
                    logger.warning(f"Missing key {e} in url_template for {self.name}")
                    # Fallback to internal ref if URL template fails
                    target = ref
            else:
                # No URL template, fall back to internal
                target = ref
        else:
            # Internal mode
            target = ref

        return {
            "target": target,
            "display_text": display_text
        }

    def get_link(
        self,
        key_value: Any,
        mode: str = "external",
    ) -> Optional[dict]:
        """
        Resolve link for a key.

        mode: "external" (default) or "internal"
        Returns dict with "target" and "display_text" or None if missing.
        """
        record = self._records.get(str(key_value))
        if not record:
            return None

        if mode == "external":
            target = record.get("url") or record.get("ref")
        else:
            target = record.get("ref")

        if not target:
            return None

        return {
            "target": target,
            "display_text": record["display_text"],
        }

    @property
    def max_display_width(self) -> int:
        """
        Maximum display text width observed across sampled links.

        Samples first 100 records, capped at 50 characters to handle outliers.
        Used for automatic column width sizing.
        """
        return self._max_display_width


class LinkedExcelWriter(ExcelWriter):
    """
    Advanced Excel writer with internal and external hyperlink management.

    LinkedExcelWriter extends ExcelWriter to enable rich, bidirectional hyperlinking
    within Excel workbooks and to external systems. It automatically caches source
    records as they're written and creates formatted hyperlinks in detail sheets that
    reference those sources.

    This is particularly powerful for creating navigable multi-sheet reports with
    master-detail relationships, drill-through capabilities, and integration with
    external CRM, ticketing, or web applications.

    Key Features
    ------------
    * **Internal navigation** - Links between worksheets (e.g., ``#Students!B5``)
    * **External integration** - Deep links to web applications
    * **Hybrid linking** - Store both internal and external, choose which to display
    * **Template-based formatting** - Use Python format strings for link text and URLs
    * **Automatic caching** - Source records cached as written, no manual tracking
    * **Mode control** - Force internal or external links per column via ``source:internal``

    Workflow
    --------
    1. Create LinkSource definitions for each linkable entity
    2. Register them with LinkedExcelWriter
    3. Write source sheets first (e.g., Students, Products)
    4. Write detail sheets with link specifications (e.g., Enrollments, Orders)
    5. Links are resolved from cache and applied automatically

    Parameters
    ----------
    file : str or Path
        Output Excel file (.xlsx)
    data : Iterable[RecordLike], optional
        Initial data to write. If None, use write_batch() for streaming mode.
    sheet_name : str, optional
        Default sheet name for write_batch() calls
    write_headers : bool, default True
        Whether to write column headers

    Examples
    --------
    **Basic internal linking between sheets**::

        with LinkedExcelWriter(file='school_report.xlsx') as writer:
            # Define linkable entity
            student_link = LinkSource(
                name="student",
                source_sheet="Students",
                key_column="student_id"
            )
            writer.register_link_source(student_link)

            # Write source sheet
            writer.write_batch(students_data, sheet_name="Students")

            # Write detail sheet with internal links
            writer.write_batch(
                enrollments_data,
                sheet_name="Enrollments",
                links={"student_name": "student:internal"}
            )

    **External links to CRM system**::

        with LinkedExcelWriter(file='sales_report.xlsx') as writer:
            customer_link = LinkSource(
                name="customer",
                source_sheet="Customers",
                key_column="customer_id",
                url_template="https://crm.company.com/customers/{crm_id}",
                text_template="{company_name} ({customer_id})"
            )
            writer.register_link_source(customer_link)

            writer.write_batch(customers_data, sheet_name="Customers")
            writer.write_batch(
                orders_data,
                sheet_name="Orders",
                links={"customer": "customer"}  # Uses external URL
            )

    **Hybrid mode with internal and external links**::

        with LinkedExcelWriter(file='support_tickets.xlsx') as writer:
            ticket_link = LinkSource(
                name="ticket",
                source_sheet="Tickets",
                key_column="ticket_id",
                url_template="https://support.company.com/ticket/{ticket_id}",
                text_template="#{ticket_id} - {subject}"
            )
            writer.register_link_source(ticket_link)

            writer.write_batch(tickets_data, sheet_name="Tickets")
            writer.write_batch(
                comments_data,
                sheet_name="Comments",
                links={
                    "ticket_link": "ticket",           # External to support system
                    "ticket_ref": "ticket:internal"    # Internal sheet navigation
                }
            )

    **Multiple link sources in one sheet**::

        with LinkedExcelWriter(file='class_roster.xlsx') as writer:
            student_link = LinkSource(
                name="student",
                source_sheet="Students",
                key_column="student_id",
                text_template="{last_name}, {first_name}"
            )
            course_link = LinkSource(
                name="course",
                source_sheet="Courses",
                key_column="course_id",
                text_template="{course_code} - {title}"
            )

            writer.register_link_source(student_link)
            writer.register_link_source(course_link)

            writer.write_batch(students_data, sheet_name="Students")
            writer.write_batch(courses_data, sheet_name="Courses")
            writer.write_batch(
                enrollments_data,
                sheet_name="Enrollments",
                links={
                    "student_name": "student:internal",
                    "course_name": "course:internal"
                }
            )

    Notes
    -----
    * Source sheets MUST be written before detail sheets that reference them
    * The key_column must exist in both source and detail datasets
    * Missing links display missing_text if set, otherwise show raw value
    * Link mode syntax: ``"source_name"`` (external) or ``"source_name:internal"``
    * Templates use str.format_map() - all fields must exist in record data
    * Hyperlink styling (blue, underlined) is applied automatically

    See Also
    --------
    LinkSource : Link definition class
    ExcelWriter : Base writer without linking capabilities
    """

    def __init__(
        self,
        data: Optional[Iterable[RecordLike]] = None,
        file: Optional[Union[str, Path]] = None,
        sheet_name: Optional[str] = None,
        headers: Optional[List[str]] = None,
        write_headers: bool = True,
        formatting: Optional[Dict] = None,
    ):
        super().__init__(data=data, file=file, sheet_name=sheet_name, headers=headers,
                         write_headers=write_headers, formatting=formatting)
        #: Registered LinkSource instances, keyed by name.
        self.link_sources: Dict[str, LinkSource] = {}

    def register_link_source(self, source: LinkSource) -> None:
        """Register a link source for use across sheets."""
        if source.name in self.link_sources:
            logger.warning(f"LinkSource '{source.name}' already registered — overwriting")
        self.link_sources[source.name] = source

    def write_batch(
        self,
        data: Iterable[RecordLike],
        sheet_name: Optional[str] = None,
        links: Optional[Dict[str, str]] = None,
    ) -> None:
        """
        Write a batch with optional hyperlinking.

        If this is the first write to this sheet in the current session, the sheet
        is cleared first. Subsequent writes to the same sheet append data.

        links: dict column_name → "source_name" or "source_name:internal"
        """
        target_sheet = sheet_name or self.active_sheet or 'Data'
        if sheet_name:
            self.active_sheet = target_sheet

        worksheet = self._get_or_create_worksheet(target_sheet)

        # Clear sheet if this is the first write to it in this session
        if target_sheet not in self._sheets_written_this_session:
            self._clear_worksheet(worksheet)
            self._sheets_written_this_session.add(target_sheet)

        # Parse links dict into resolved mapping: column → (source, mode)
        link_mapping = {}
        if links:
            for col, spec in links.items():
                if ':' in spec:
                    source_name, mode_str = spec.split(':', 1)
                    mode = mode_str.lower()
                    if mode not in {'internal', 'external'}:
                        raise ValueError(f"Invalid mode '{mode_str}' in link spec '{spec}'")
                else:
                    source_name = spec
                    mode = "external"

                if source_name not in self.link_sources:
                    raise ValueError(f"Unknown LinkSource '{source_name}'")

                link_mapping[col] = (self.link_sources[source_name], mode)

        # Determine if this sheet is a source sheet for any registered LinkSource
        # Skip external_only sources since they don't cache
        source_for_this_sheet = [
            src for src in self.link_sources.values()
            if not src.external_only and src.source_sheet == target_sheet
        ]

        row_count = self._write_to_worksheet(
            data=data,
            worksheet=worksheet,
            write_headers=self.write_headers,
            link_mapping=link_mapping,
            source_for_this_sheet=source_for_this_sheet,
            target_sheet=target_sheet
        )

        self._row_num += row_count
        logger.info(f"Wrote {row_count} rows to sheet '{target_sheet}' with linking")

    def _write_to_worksheet(
        self,
        data: Iterable[RecordLike],
        worksheet: 'Worksheet',
        columns: Optional[List[str]] = None,
        write_headers: bool = True,
        link_mapping: Optional[Dict[str, tuple]] = None,
        source_for_this_sheet: Optional[list] = None,
        target_sheet: Optional[str] = None
    ) -> int:
        link_mapping = link_mapping or {}
        source_for_this_sheet = source_for_this_sheet or []

        # Lazy init columns - always update columns for each batch
        self.data_iterator, self.columns = self._get_data_iterator(data, columns)

        if not self.columns:
            raise ValueError("Could not determine columns from data")

        col_fmt = self._build_col_fmt_map(self.columns)
        rows_fmt = self.formatting.get('rows', {})
        row_style_fn = rows_fmt.get('style') if isinstance(rows_fmt, dict) else None
        odd_style = rows_fmt.get('odd', {}).get('format') if isinstance(rows_fmt, dict) else None
        even_style = rows_fmt.get('even', {}).get('format') if isinstance(rows_fmt, dict) else None

        row_count = 0
        header_widths = [len(col) for col in self.columns]
        data_widths = [0] * len(self.columns)
        width_sample_size = 15
        header_font = Font(bold=True)

        should_write_headers = write_headers and (worksheet.cell(1, 1).value is None)
        data_start_row = 2 if should_write_headers else worksheet.max_row + 1

        if should_write_headers:
            for col_idx, column_name in enumerate(self.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=column_name)
                hfmt = col_fmt[col_idx - 1].get('header_format') if col_fmt else None
                if hfmt:
                    cell.style = hfmt
                else:
                    cell.font = header_font

        col_index_map = {name: idx + 1 for idx, name in enumerate(self.columns)}

        for row_idx, record in enumerate(self.data_iterator, data_start_row):
            row_dict = record
            values = self._row_to_tuple(record)
            data_row_num = row_idx - data_start_row + 1
            alt_style = odd_style if data_row_num % 2 else even_style
            row_style = row_style_fn(record) if callable(row_style_fn) else alt_style

            if isinstance(rows_fmt, dict) and data_row_num in rows_fmt:
                rp = rows_fmt[data_row_num]
                if isinstance(rp, dict) and 'height' in rp:
                    worksheet.row_dimensions[row_idx].height = rp['height']

            for col_idx, value in enumerate(values, 1):
                col_name = self.columns[col_idx - 1]
                cell = worksheet.cell(row=row_idx, column=col_idx)
                is_date_val = isinstance(value, (date, datetime))

                # Set cell value (handles dates, None, type preservation)
                if isinstance(value, datetime) and value.time() != MIDNIGHT:
                    cell.value = value
                    cell.style = 'datetime_style'
                    if row_count < width_sample_size:
                        data_widths[col_idx - 1] = max(data_widths[col_idx - 1], 19)
                elif isinstance(value, (date, datetime)):
                    cell.value = value
                    cell.style = 'date_style'
                    if row_count < width_sample_size:
                        data_widths[col_idx - 1] = max(data_widths[col_idx - 1], 10)
                elif value is None:
                    cell.value = ''
                else:
                    cell.value = value
                    if row_count < width_sample_size:
                        data_widths[col_idx - 1] = max(data_widths[col_idx - 1], len(str(value)))

                # Apply column / row formatting (skipped for date cells)
                if not is_date_val and col_fmt:
                    col_style = col_fmt[col_idx - 1].get('format')
                    if col_style:
                        cell.style = col_style

                if row_style and not is_date_val:
                    cell.style = row_style

                # Link spec overrides everything (hyperlink_style takes precedence)
                link_spec = link_mapping.get(col_name)
                if link_spec:
                    source, mode = link_spec

                    if source.external_only:
                        link_info = source.generate_link_from_row(row_dict, ref="", mode="external", column_value=value)
                    elif source.source_sheet == target_sheet:
                        key_col_letter = get_column_letter(col_index_map[source.key_column])
                        ref = f"#{target_sheet}!{key_col_letter}{row_idx}"
                        link_info = source.generate_link_from_row(row_dict, ref, mode=mode)
                    else:
                        key_value = value
                        link_info = source.get_link(key_value, mode=mode) if key_value is not None else None

                    if link_info:
                        cell.hyperlink = link_info["target"]
                        cell.value = link_info["display_text"]
                        cell.style = 'hyperlink_style'
                    elif source.missing_text is not None:
                        cell.value = source.missing_text

            # Cache row for future cross-sheet linking
            for source in source_for_this_sheet:
                key_col_letter = get_column_letter(col_index_map[source.key_column])
                ref = f"#{target_sheet}!{key_col_letter}{row_idx}"
                key_value = row_dict.get(source.key_column)
                if key_value is not None:
                    source.cache_record(key_value, row_dict, ref)

            row_count += 1

        if should_write_headers:
            self._finalize_headers(worksheet, header_widths, data_widths, col_fmt, rows_fmt,
                                   link_mapping=link_mapping)

        return row_count


def check_dependencies():
    if not HAS_OPENPYXL:
        logger.error('Openpyxl is not available. Excel files not supported.')

check_dependencies()