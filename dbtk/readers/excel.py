# dbtk/readers/excel.py
import datetime as dt
import logging
from typing import List, Any, Iterator, Optional
from .base import Reader, Clean, ReturnType

logger = logging.getLogger(__name__)

# Check for optional dependencies
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning('openpyxl not available. xlsx files not supported.')

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    if not HAS_OPENPYXL:
        logger.warning('xlrd not available. xls files not supported.')


class XLSXReader(Reader):
    """Class to iterate over an Excel Spreadsheet using openpyxl."""

    def __init__(self,
                 worksheet,
                 headers: Optional[List[str]] = None,
                 add_rownum: bool = True,
                 clean_headers: Clean = Clean.DEFAULT,
                 skip_records: int = 0,
                 max_records: Optional[int] = None,
                 return_type: str = ReturnType.DEFAULT):
        """Initialize XLSXReader for reading Excel .xlsx files.

        Args:
            worksheet: openpyxl.Worksheet object to read from.
            headers: Optional list of header names to use instead of reading from row 1.
            add_rownum: If True, adds a rownum field to each record (default: True).
            clean_headers: Header cleaning level from Clean enum (default: Clean.DEFAULT).
            skip_records: Number of data records to skip after headers (default: 0).
            max_records: Maximum number of records to read, or None for all (default: None).
            return_type: Either 'record' for Record objects or 'dict' for OrderedDict.

        Raises:
            TypeError: If worksheet is not an openpyxl.Worksheet.
        """
        if worksheet.__class__.__name__ != 'Worksheet':
            raise TypeError('worksheet must be of type openpyxl.Worksheet or use XLReader')
        super().__init__(add_rownum=add_rownum, clean_headers=clean_headers,
                         skip_records=skip_records, max_records=max_records,
                         return_type=return_type)
        self.ws = worksheet
        self._headers_read = False
        self._raw_headers = headers  # Use provided headers if given
        self._start_row = 1 if headers else 2  # openpyxl 1-based indexing

    def _read_headers(self) -> List[str]:
        """Read the header row from the Excel worksheet (row 1) or use provided headers.

        Returns:
            List of header values.

        Raises:
            StopIteration: If the file is empty and no headers are provided.
        """
        if self._raw_headers is not None:
            return self._raw_headers
        if not self._headers_read:
            # openpyxl is 1-based, so row 1 is the first row
            if self.ws.max_row < 1:
                raise StopIteration("Empty worksheet")
            header_cells = self.ws[1]
            self._raw_headers = [cell.value for cell in header_cells]
            self._headers_read = True
        return self._raw_headers

    def _generate_rows(self) -> Iterator[List[Any]]:
        """Yield data rows from the Excel worksheet starting from _start_row.

        Yields:
            List of cell values for each data row.
        """
        for row_num in range(self._start_row, self.ws.max_row + 1):  # 1-based
            row_cells = self.ws[row_num]
            yield [cell.value for cell in row_cells]

    def _cleanup(self):
        """Perform cleanup (no-op for openpyxl worksheet as no file pointer is managed)."""
        pass


class XLReader(Reader):
    """Class to iterate over an Excel Spreadsheet using xlrd."""

    def __init__(self,
                 worksheet,
                 headers: Optional[List[str]] = None,
                 add_rownum: bool = True,
                 clean_headers: Clean = Clean.DEFAULT,
                 skip_records: int = 0,
                 max_records: Optional[int] = None,
                 return_type: str = ReturnType.DEFAULT):
        """Initialize XLReader for reading Excel .xls files.

        Args:
            worksheet: xlrd.Sheet object to read from.
            headers: Optional list of header names to use instead of reading from row 0.
            add_rownum: If True, adds a rownum field to each record (default: True).
            clean_headers: Header cleaning level from Clean enum (default: Clean.DEFAULT).
            skip_records: Number of data records to skip after headers (default: 0).
            max_records: Maximum number of records to read, or None for all (default: None).
            return_type: Either 'record' for Record objects or 'dict' for OrderedDict.

        Raises:
            TypeError: If worksheet is not an xlrd.Sheet.
        """
        if worksheet.__class__.__name__ != 'Sheet':
            raise TypeError('worksheet must be of type xlrd.Sheet or use XLSXReader')
        super().__init__(add_rownum=add_rownum, clean_headers=clean_headers,
                         skip_records=skip_records, max_records=max_records,
                         return_type=return_type)
        self.ws = worksheet
        self.datemode = worksheet.book.datemode
        self._headers_read = False
        self._raw_headers = headers  # Use provided headers if given
        self._start_row = 1 if headers is not None else 0  # xlrd 0-based indexing

    def _read_headers(self) -> List[str]:
        """Read the header row from the Excel worksheet (row 0) or use provided headers.

        Returns:
            List of header values.

        Raises:
            StopIteration: If the file is empty and no headers are provided.
        """
        if self._raw_headers is not None:
            return self._raw_headers
        if not self._headers_read:
            if self.ws.nrows < 1:
                raise StopIteration("Empty worksheet")
            header_cells = self.ws.row(0)  # xlrd is 0-based
            self._raw_headers = [cell.value for cell in header_cells]
            self._headers_read = True
        return self._raw_headers

    def _generate_rows(self) -> Iterator[List[Any]]:
        """Yield data rows from the Excel worksheet starting from _start_row.

        Yields:
            List of converted cell values for each data row.
        """
        for row_num in range(self._start_row, self.ws.nrows):  # 0-based
            row_cells = self.ws.row(row_num)
            yield [self._convert_cell_value(cell) for cell in row_cells]

    def _convert_cell_value(self, cell) -> Any:
        """Convert an Excel cell value to an appropriate Python type.

        Args:
            cell: xlrd cell object to convert.

        Returns:
            Converted value (datetime, int, float, str, or None).
        """
        if cell is None or cell.ctype == 0:
            return None
        elif cell.ctype == 3:
            try:
                tm_tuple = xlrd.xldate_as_tuple(cell.value, self.datemode)
                if any(tm_tuple[:3]):
                    return dt.datetime(*tm_tuple)
                else:
                    return dt.time(*tm_tuple[3:])
            except (ValueError, TypeError):
                return None
        elif cell.ctype == 2:
            if cell.value == int(cell.value):
                return int(cell.value)
            else:
                return cell.value
        else:
            try:
                return str(cell.value)
            except (ValueError, TypeError):
                return cell.value

    def _cleanup(self):
        """Perform cleanup (no-op for xlrd worksheet as no file pointer is managed)."""
        pass


def open_workbook(filename: str):
    """Open an Excel workbook using openpyxl for .xlsx or xlrd for .xls.

    Args:
        filename: Path to the Excel file (.xlsx or .xls).

    Returns:
        Workbook object (openpyxl.Workbook or xlrd.Book).

    Raises:
        ImportError: If neither openpyxl nor xlrd is available.
    """
    if HAS_OPENPYXL and filename.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(filename, data_only=True)
    elif HAS_XLRD:
        workbook = xlrd.open_workbook(filename)
    else:
        raise ImportError('Neither openpyxl nor xlrd available for Excel file support')
    return workbook


def get_sheet_by_index(wb, index: int):
    """Get a worksheet from a workbook by index.

    Args:
        wb: Workbook object (openpyxl.Workbook or xlrd.Book).
        index: Index of the sheet to retrieve (0-based).

    Returns:
        Worksheet object (openpyxl.Worksheet or xlrd.Sheet).

    Raises:
        TypeError: If workbook type is not supported.
    """
    if wb.__class__.__name__ == 'Workbook':
        return wb.worksheets[index]
    elif wb.__class__.__name__ == 'Book':
        return wb.sheet_by_index(index)
    else:
        raise TypeError(f"Unknown workbook type: {wb.__class__.__name__}")


def check_dependencies():
    """Check for optional dependencies and issue warnings if missing."""
    if not HAS_OPENPYXL and not HAS_XLRD:
        logger.error('Neither openpyxl nor xlrd available. Excel files not supported.')

check_dependencies()