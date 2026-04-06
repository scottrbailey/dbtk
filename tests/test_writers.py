# tests/test_writers.py
"""
Tests for dbtk writers module.
Channeling data out to various formats like a master waterbender directing flow.
"""

import pytest
import json
from pathlib import Path
from datetime import date, datetime
from collections import namedtuple

from dbtk.readers import CSVReader
from dbtk.utils import FixedColumn
from dbtk.writers import (
    CSVWriter, ExcelWriter, FixedWidthWriter,
    JSONWriter, NDJSONWriter, XMLWriter,
    to_csv, to_excel, to_fixed_width, to_json, to_ndjson, to_xml
)

# FixedColumn schema matching the sample_data fixtures (8 fields, 122 chars/line)
SAMPLE_COLUMNS = [
    FixedColumn('trainee_id',        1,   5, 'int'),
    FixedColumn('monk_name',         6,  35),
    FixedColumn('home_temple',      36,  65),
    FixedColumn('mastery_rank',     66,  70, 'int'),
    FixedColumn('bison_companion',  71,  82),
    FixedColumn('daily_meditation', 83,  90, 'float'),
    FixedColumn('birth_date',       91, 102),
    FixedColumn('last_training',   103, 122),
]
SAMPLE_LINE_WIDTH = sum(c.width for c in SAMPLE_COLUMNS)  # 122


# Fixtures
@pytest.fixture
def fixtures_dir():
    """Path to test fixtures directory."""
    return Path(__file__).parent / 'fixtures' / 'readers'


@pytest.fixture
def csv_file(fixtures_dir):
    """Path to CSV test file."""
    return fixtures_dir / 'sample_data.csv'


@pytest.fixture
def excel_file(fixtures_dir):
    """Path to Excel test file."""
    return fixtures_dir / 'sample_data.xlsx'


@pytest.fixture
def sample_records(excel_file):
    """Load first 10 records from sample_data.xlsx as Records with types preserved."""
    from dbtk.readers.excel import open_workbook, get_sheet_by_index
    from dbtk.readers import XLSXReader, XLSReader

    wb = open_workbook(str(excel_file))
    ws = get_sheet_by_index(wb, 0)

    reader_class = XLSXReader if ws.__class__.__name__ == 'Worksheet' else XLSReader

    with reader_class(ws, add_row_num=False) as reader:
        records = []
        for i, record in enumerate(reader):
            if i >= 10:
                break
            records.append(record)
    return records


@pytest.fixture
def sample_dicts(sample_records):
    """Convert sample records to list of dicts."""
    return [record.to_dict() for record in sample_records]


@pytest.fixture
def sample_namedtuples(sample_records):
    """Convert sample records to list of namedtuples."""
    if not sample_records:
        return []

    # Create namedtuple type from first record's keys
    RecordTuple = namedtuple('RecordTuple', sample_records[0].keys())
    return [RecordTuple(**record.to_dict()) for record in sample_records]


@pytest.fixture
def sample_lists(sample_records):
    """Convert sample records to list of lists."""
    return [list(record.values()) for record in sample_records]


@pytest.fixture
def sample_columns(sample_records):
    """Get column names from sample records."""
    if sample_records:
        return list(sample_records[0].keys())
    return []


# Base Writer Tests using CSV
class TestBaseWriter:
    """Tests for base writer functionality using CSVWriter."""

    def test_write_from_records(self, tmp_path, sample_records):
        """Test writing from list of Record objects."""
        output_file = tmp_path / "output.csv"

        to_csv(sample_records, output_file)

        # Read back and verify
        with CSVReader(open(output_file, encoding='utf-8-sig')) as reader:
            records = list(reader)
            assert len(records) == 10
            assert records[0]['trainee_id'] == '1'

    def test_write_from_dicts(self, tmp_path, sample_dicts):
        """Test writing from list of dictionaries."""
        output_file = tmp_path / "output.csv"

        to_csv(sample_dicts, output_file)

        # Read back and verify
        with CSVReader(open(output_file, encoding='utf-8-sig')) as reader:
            records = list(reader)
            assert len(records) == 10
            assert 'trainee_id' in records[0]

    def test_write_from_namedtuples(self, tmp_path, sample_namedtuples):
        """Test writing from list of namedtuples."""
        output_file = tmp_path / "output.csv"

        to_csv(sample_namedtuples, output_file)

        # Read back and verify
        with CSVReader(open(output_file, encoding='utf-8-sig')) as reader:
            records = list(reader)
            assert len(records) == 10

    def test_write_from_lists(self, tmp_path, sample_lists, sample_columns):
        """Test writing from list of lists with explicit columns."""
        output_file = tmp_path / "output.csv"

        with CSVWriter(sample_lists, output_file, columns=sample_columns) as writer:
            writer.write()

        # Read back and verify
        with CSVReader(open(output_file, encoding='utf-8-sig'), add_row_num=False) as reader:
            records = list(reader)
            assert len(records) == 10
            assert reader.headers == sample_columns

    def test_write_headers_true(self, tmp_path, sample_records):
        """Test write_headers=True writes header row."""
        output_file = tmp_path / "output.csv"

        to_csv(sample_records, output_file, write_headers=True)

        with open(output_file, encoding='utf-8-sig') as f:
            first_line = f.readline().strip()
            assert 'trainee_id' in first_line

    def test_write_headers_false(self, tmp_path, sample_records):
        """Test write_headers=False omits header row."""
        output_file = tmp_path / "output.csv"

        to_csv(sample_records, output_file, write_headers=False)

        with open(output_file, encoding='utf-8-sig') as f:
            first_line = f.readline().strip()
            # First line should be data, not headers
            assert first_line.startswith('1,') or first_line.startswith('"1"')

    def test_to_string_date(self, tmp_path, sample_records):
        """Test BaseWriter.to_string() handles dates correctly."""
        output_file = tmp_path / "output.csv"
        writer = CSVWriter(sample_records, output_file)

        test_date = date(2024, 12, 25)
        result = writer.to_string(test_date)
        assert result == '2024-12-25'

    def test_to_string_datetime_no_microseconds(self, tmp_path, sample_records):
        """Test BaseWriter.to_string() handles datetime without microseconds."""
        output_file = tmp_path / "output.csv"
        writer = CSVWriter(sample_records, output_file)

        test_datetime = datetime(2024, 12, 25, 15, 30, 45)
        result = writer.to_string(test_datetime)
        assert result == '2024-12-25 15:30:45'

    def test_to_string_datetime_with_microseconds(self, tmp_path, sample_records):
        """Test BaseWriter.to_string() handles datetime with microseconds."""
        output_file = tmp_path / "output.csv"
        writer = CSVWriter(sample_records, output_file)

        test_datetime = datetime(2024, 12, 25, 15, 30, 45, 123456)
        result = writer.to_string(test_datetime)
        assert result == '2024-12-25 15:30:45.123456'

    def test_to_string_datetime_at_midnight(self, tmp_path, sample_records):
        """Test BaseWriter.to_string() handles datetime at midnight as date."""
        output_file = tmp_path / "output.csv"
        writer = CSVWriter(sample_records, output_file)

        test_datetime = datetime(2024, 12, 25, 0, 0, 0)
        result = writer.to_string(test_datetime)
        assert result == '2024-12-25'

    def test_to_string_none(self, tmp_path, sample_records):
        """Test BaseWriter.to_string() handles None."""
        output_file = tmp_path / "output.csv"
        writer = CSVWriter(sample_records, output_file)

        result = writer.to_string(None)
        assert result == ''

    def test_to_string_number(self, tmp_path, sample_records):
        """Test BaseWriter.to_string() handles numbers."""
        output_file = tmp_path / "output.csv"
        writer = CSVWriter(sample_records, output_file)

        assert writer.to_string(42) == '42'
        assert writer.to_string(3.14) == '3.14'

    def test_row_count(self, tmp_path, sample_records):
        """Test that row_count is correctly tracked."""
        output_file = tmp_path / "output.csv"

        writer = CSVWriter(sample_records, output_file)
        count = writer.write()

        assert count == 10
        assert writer.row_count == 10


class TestCSVWriter:
    """Tests specific to CSV writer."""

    def test_custom_delimiter_tab(self, tmp_path, sample_records):
        """Test writing with tab delimiter."""
        output_file = tmp_path / "output.tsv"

        to_csv(sample_records, output_file, delimiter='\t')

        with open(output_file, encoding='utf-8-sig') as f:
            first_line = f.readline()
            assert '\t' in first_line

    def test_custom_delimiter_pipe(self, tmp_path, sample_records):
        """Test writing with pipe delimiter."""
        output_file = tmp_path / "output.csv"

        to_csv(sample_records, output_file, delimiter='|')

        with open(output_file, encoding='utf-8-sig') as f:
            first_line = f.readline()
            assert '|' in first_line

    def test_custom_quotechar(self, tmp_path, sample_records):
        """Test writing with custom quote character."""
        output_file = tmp_path / "output.csv"

        to_csv(sample_records, output_file, quotechar="'")

        # Just verify file was written successfully
        assert output_file.exists()

    def test_round_trip_preserves_data(self, tmp_path, sample_records):
        """Test that writing and reading back preserves data."""
        output_file = tmp_path / "output.csv"

        # Write
        to_csv(sample_records, output_file)

        # Read back
        with CSVReader(open(output_file, encoding='utf-8-sig')) as reader:
            records_back = list(reader)

        # Compare (values will be strings after CSV round-trip)
        assert len(records_back) == len(sample_records)
        assert str(sample_records[0]['trainee_id']) == records_back[0]['trainee_id']
        assert sample_records[0]['monk_name'] == records_back[0]['monk_name']

    def test_csv_write_batch_multiple_calls(self, tmp_path, sample_records):
        """Test CSVWriter.write_batch() called multiple times."""
        output_file = tmp_path / "output.csv"

        with CSVWriter(data=None, file=output_file) as writer:
            # Write in batches
            writer.write_batch(sample_records[:5])
            writer.write_batch(sample_records[5:])

        # Read back
        with CSVReader(open(output_file, encoding='utf-8-sig')) as reader:
            records_back = list(reader)

        assert len(records_back) == 10

    def test_csv_write_batch_headers_only_once(self, tmp_path, sample_records):
        """Test CSV headers written only on first batch."""
        output_file = tmp_path / "output.csv"

        with CSVWriter(data=None, file=output_file) as writer:
            writer.write_batch(sample_records[:3])
            writer.write_batch(sample_records[3:6])
            writer.write_batch(sample_records[6:])

        # Check file content
        with open(output_file, encoding='utf-8-sig') as f:
            lines = f.readlines()

        # Should have 1 header line + 10 data lines
        assert len(lines) == 11
        assert 'trainee_id' in lines[0]

    def test_csv_write_batch_streaming(self, tmp_path, sample_records):
        """Test CSV streaming with batches."""
        output_file = tmp_path / "output.csv"

        # Open file handle and use streaming mode
        with open(output_file, 'w', encoding='utf-8-sig', newline='') as fp:
            writer = CSVWriter(data=None, file=fp)
            for i in range(0, len(sample_records), 2):
                batch = sample_records[i:i+2]
                writer.write_batch(batch)

        # Read back and verify
        with CSVReader(open(output_file, encoding='utf-8-sig')) as reader:
            records_back = list(reader)

        assert len(records_back) == 10


class TestExcelWriter:
    """Tests specific to Excel writer."""

    pytestmark = pytest.mark.skipif(
        __import__("importlib").util.find_spec("openpyxl") is None,
        reason="openpyxl not installed",
    )

    def test_write_to_excel(self, tmp_path, sample_records):
        """Test basic Excel writing."""
        output_file = tmp_path / "output.xlsx"

        to_excel(sample_records, output_file)

        assert output_file.exists()

        # Read back with openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active

        # Check headers
        assert ws.cell(1, 1).value == 'trainee_id'

        # Check first data row
        assert ws.cell(2, 1).value == 1

    def test_custom_sheet_name(self, tmp_path, sample_records):
        """Test writing to custom sheet name."""
        output_file = tmp_path / "output.xlsx"

        to_excel(sample_records, output_file, sheet='Trainees')

        from openpyxl import load_workbook
        wb = load_workbook(output_file)

        assert 'Trainees' in wb.sheetnames

    def test_multiple_sheets(self, tmp_path, sample_records):
        """Test writing to multiple sheets in same workbook."""
        output_file = tmp_path / "output.xlsx"

        # Write first sheet
        to_excel(sample_records[:5], output_file, sheet='First')

        # Write second sheet
        to_excel(sample_records[5:], output_file, sheet='Second')

        from openpyxl import load_workbook
        wb = load_workbook(output_file)

        assert 'First' in wb.sheetnames
        assert 'Second' in wb.sheetnames

    def test_date_type_preservation(self, tmp_path, sample_records):
        """Test that dates are written as date types in Excel."""
        output_file = tmp_path / "output.xlsx"

        to_excel(sample_records, output_file)

        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active

        # Find birth_date column (column 7)
        birth_date_cell = ws.cell(2, 7)

        # Should be a datetime object, not string
        assert isinstance(birth_date_cell.value, datetime)

    def test_write_batch_multiple_calls_same_sheet(self, tmp_path, sample_records):
        """Test calling write_batch multiple times appends to same sheet."""
        output_file = tmp_path / "output.xlsx"

        with ExcelWriter(file=output_file) as writer:
            # Write first batch
            writer.write_batch(sample_records[:5], sheet_name='Data')
            # Write second batch
            writer.write_batch(sample_records[5:], sheet_name='Data')

        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb['Data']

        # Should have headers + 10 data rows
        assert ws.max_row == 11
        assert ws.cell(1, 1).value == 'trainee_id'
        assert ws.cell(2, 1).value == 1
        assert ws.cell(11, 1).value == 10

    def test_write_batch_multiple_sheets(self, tmp_path, sample_records):
        """Test write_batch to multiple different sheets in one workbook."""
        output_file = tmp_path / "output.xlsx"

        with ExcelWriter(file=output_file) as writer:
            writer.write_batch(sample_records[:5], sheet_name='First')
            writer.write_batch(sample_records[5:], sheet_name='Second')

        from openpyxl import load_workbook
        wb = load_workbook(output_file)

        assert 'First' in wb.sheetnames
        assert 'Second' in wb.sheetnames

        # Check row counts
        assert wb['First'].max_row == 6  # headers + 5 data
        assert wb['Second'].max_row == 6  # headers + 5 data

    def test_write_batch_headers_only_once(self, tmp_path, sample_records):
        """Test that headers are written only on first batch, not subsequent."""
        output_file = tmp_path / "output.xlsx"

        with ExcelWriter(file=output_file) as writer:
            writer.write_batch(sample_records[:3], sheet_name='Data')
            writer.write_batch(sample_records[3:6], sheet_name='Data')
            writer.write_batch(sample_records[6:], sheet_name='Data')

        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb['Data']

        # Should have exactly 1 header row + 10 data rows
        assert ws.max_row == 11

        # Check first row is headers
        assert ws.cell(1, 1).value == 'trainee_id'

        # Check second row is data (not duplicate headers)
        assert ws.cell(2, 1).value == 1

    def test_write_batch_streaming_mode(self, tmp_path, sample_records):
        """Test streaming mode with write_batch (data=None in init)."""
        output_file = tmp_path / "output.xlsx"

        # Simulate streaming: no data in __init__
        with ExcelWriter(file=output_file) as writer:
            # Write in batches
            for i in range(0, len(sample_records), 3):
                batch = sample_records[i:i+3]
                writer.write_batch(batch, sheet_name='Stream')

        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb['Stream']

        # Should have all records
        assert ws.max_row == 11  # headers + 10 data rows

    def test_write_batch_append_to_existing_workbook(self, tmp_path, sample_records):
        """Test appending new sheet to existing workbook."""
        output_file = tmp_path / "output.xlsx"

        # First session: create workbook with one sheet
        with ExcelWriter(file=output_file) as writer:
            writer.write_batch(sample_records[:5], sheet_name='FirstRun')

        # Second session: append another sheet
        with ExcelWriter(file=output_file) as writer:
            writer.write_batch(sample_records[5:], sheet_name='SecondRun')

        from openpyxl import load_workbook
        wb = load_workbook(output_file)

        assert 'FirstRun' in wb.sheetnames
        assert 'SecondRun' in wb.sheetnames
        assert wb['FirstRun'].max_row == 6
        assert wb['SecondRun'].max_row == 6

    def test_write_batch_default_sheet_name(self, tmp_path, sample_records):
        """Test that default sheet name 'Data' is used when not specified."""
        output_file = tmp_path / "output.xlsx"

        with ExcelWriter(file=output_file) as writer:
            writer.write_batch(sample_records)

        from openpyxl import load_workbook
        wb = load_workbook(output_file)

        assert 'Data' in wb.sheetnames

    def test_write_batch_custom_active_sheet(self, tmp_path, sample_records):
        """Test using sheet_name parameter in __init__ as default."""
        output_file = tmp_path / "output.xlsx"

        with ExcelWriter(file=output_file, sheet_name='MySheet') as writer:
            # Don't specify sheet_name - should use 'MySheet'
            writer.write_batch(sample_records)

        from openpyxl import load_workbook
        wb = load_workbook(output_file)

        assert 'MySheet' in wb.sheetnames
        assert wb['MySheet'].max_row == 11

    def test_headers_parameter_override(self, tmp_path):
        """Test that headers parameter overrides detected field names in header row."""
        from dbtk.record import Record

        # Create records with lowercase field names (simulating cursor normalization)
        RecordClass = type('TestRecord', (Record,), {})
        RecordClass.set_fields(['user_id', 'user_name', 'email_address'])

        records = [
            RecordClass(1, 'Alice', 'alice@example.com'),
            RecordClass(2, 'Bob', 'bob@example.com'),
        ]

        output_file = tmp_path / "output.xlsx"

        # Write with custom header names (original database column names)
        original_headers = ['USER_ID', 'USER_NAME', 'EMAIL_ADDRESS']
        to_excel(records, output_file, headers=original_headers)

        # Read back and verify headers
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active

        # Check that headers match the provided headers, not the Record field names
        assert ws.cell(1, 1).value == 'USER_ID'
        assert ws.cell(1, 2).value == 'USER_NAME'
        assert ws.cell(1, 3).value == 'EMAIL_ADDRESS'

        # Check data values are still extracted correctly from Record fields
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == 'Alice'
        assert ws.cell(2, 3).value == 'alice@example.com'


@pytest.fixture
def fw_records(csv_file):
    """Load first 10 records from CSV (no Excel dependency) for FixedWidth tests."""
    with CSVReader(open(csv_file), add_row_num=False) as reader:
        return [r for i, r in enumerate(reader) if i < 10]


class TestFixedWidthWriter:
    """Tests specific to fixed-width writer."""

    def test_write_fixed_width(self, tmp_path, fw_records):
        """Basic write: each line is exactly the sum of column widths."""
        output_file = tmp_path / "output.txt"
        to_fixed_width(fw_records, SAMPLE_COLUMNS, output_file)
        assert output_file.exists()
        with open(output_file, encoding='utf-8-sig') as f:
            first_line = f.readline()
        assert len(first_line.rstrip('\n')) == SAMPLE_LINE_WIDTH

    def test_int_fields_right_aligned(self, tmp_path, fw_records):
        """Integer columns are right-aligned with leading spaces (from FixedColumn type)."""
        output_file = tmp_path / "output.txt"
        to_fixed_width(fw_records, SAMPLE_COLUMNS, output_file)
        with open(output_file, encoding='utf-8-sig') as f:
            first_line = f.readline()
        trainee_id = first_line[:5]     # width-5 int column
        assert trainee_id.strip('0') == '1'
        assert trainee_id.startswith('0')   # int column: right-aligned, zero-padded

    def test_text_fields_left_aligned(self, tmp_path, fw_records):
        """Text columns are left-aligned with trailing spaces."""
        output_file = tmp_path / "output.txt"
        to_fixed_width(fw_records, SAMPLE_COLUMNS, output_file)
        with open(output_file, encoding='utf-8-sig') as f:
            first_line = f.readline()
        monk_name = first_line[5:35]    # width-30 text column
        assert monk_name.rstrip() == 'Master Aang'
        assert monk_name.endswith(' ')  # left-aligned → trailing spaces

    def test_truncate_overflow_true(self, tmp_path, fw_records):
        """truncate_overflow=True silently clips values that exceed column width."""
        narrow = [
            FixedColumn('trainee_id',        1,   5, 'int'),
            FixedColumn('monk_name',         6,  10),           # too narrow for long names
            FixedColumn('home_temple',      11,  15),
            FixedColumn('mastery_rank',     16,  20, 'int'),
            FixedColumn('bison_companion',  21,  32),
            FixedColumn('daily_meditation', 33,  40, 'float'),
            FixedColumn('birth_date',       41,  52),
            FixedColumn('last_training',    53,  72),
        ]
        output_file = tmp_path / "output.txt"
        to_fixed_width(fw_records, narrow, output_file, truncate_overflow=True)
        assert output_file.exists()

    def test_truncate_overflow_false(self, tmp_path, fw_records):
        """truncate_overflow=False raises ValueError on overflow."""
        narrow = [
            FixedColumn('trainee_id',        1,   5, 'int'),
            FixedColumn('monk_name',         6,  10),           # too narrow
            FixedColumn('home_temple',      11,  15),
            FixedColumn('mastery_rank',     16,  20, 'int'),
            FixedColumn('bison_companion',  21,  32),
            FixedColumn('daily_meditation', 33,  40, 'float'),
            FixedColumn('birth_date',       41,  52),
            FixedColumn('last_training',    53,  72),
        ]
        output_file = tmp_path / "output.txt"
        with pytest.raises(ValueError, match="too large"):
            to_fixed_width(fw_records, narrow, output_file, truncate_overflow=False)

    def test_fixed_width_record_passthrough(self, tmp_path):
        """FixedWidthRecord input goes straight to to_line() without re-casting."""
        from dbtk.record import FixedWidthRecord
        cols = [FixedColumn('code', 1, 2), FixedColumn('amount', 3, 12, 'int', pad_char='0')]
        cls = type('R', (FixedWidthRecord,), {})
        cls.set_fields(cols)
        records = [cls('AB', 42), cls('CD', 999)]

        output_file = tmp_path / "output.txt"
        to_fixed_width(records, cols, output_file)

        lines = output_file.read_text().splitlines()
        assert lines[0] == 'AB0000000042'
        assert lines[1] == 'CD0000000999'

    def test_columns_required(self, tmp_path, fw_records):
        """Omitting columns raises ValueError."""
        with pytest.raises(ValueError, match="columns"):
            FixedWidthWriter(fw_records, tmp_path / "out.txt")


class TestJSONWriter:
    """Tests specific to JSON writer."""

    def test_write_json_array(self, tmp_path, sample_records):
        """Test writing JSON as array of objects."""
        output_file = tmp_path / "output.json"

        to_json(sample_records, output_file)

        assert output_file.exists()

        # Parse and verify
        with open(output_file, encoding='utf-8-sig') as f:
            data = json.load(f)

        assert isinstance(data, list)
        assert len(data) == 10
        assert data[0]['trainee_id'] == 1

    def test_json_pretty_print(self, tmp_path, sample_records):
        """Test JSON with indentation (pretty print)."""
        output_file = tmp_path / "output.json"

        to_json(sample_records, output_file, indent=2)

        with open(output_file, encoding='utf-8-sig') as f:
            content = f.read()

        # Should have newlines and indentation
        assert '\n' in content
        assert '  ' in content

    def test_json_compact(self, tmp_path, sample_records):
        """Test JSON without indentation (compact)."""
        output_file = tmp_path / "output.json"

        to_json(sample_records, output_file, indent=None)

        with open(output_file, encoding='utf-8-sig') as f:
            content = f.read()

        # Should be mostly on one line (no pretty printing)
        lines = content.strip().split('\n')
        assert len(lines) == 1

    def test_json_date_serialization(self, tmp_path, sample_records):
        """Test that dates are serialized as strings in JSON."""
        output_file = tmp_path / "output.json"

        to_json(sample_records, output_file)

        with open(output_file, encoding='utf-8-sig') as f:
            data = json.load(f)

        # Dates should be strings
        assert isinstance(data[0]['birth_date'], str)
        assert data[0]['birth_date'] == '1965-12-16'


class TestNDJSONWriter:
    """Tests specific to NDJSON writer."""

    def test_write_ndjson(self, tmp_path, sample_records):
        """Test writing NDJSON (one JSON object per line)."""
        output_file = tmp_path / "output.ndjson"

        to_ndjson(sample_records, output_file)

        assert output_file.exists()

        # Parse and verify - each line should be valid JSON
        with open(output_file, encoding='utf-8-sig') as f:
            lines = f.readlines()

        assert len(lines) == 10

        # Parse first line
        first_record = json.loads(lines[0])
        assert first_record['trainee_id'] == 1

    def test_ndjson_no_indentation(self, tmp_path, sample_records):
        """Test NDJSON has no indentation."""
        output_file = tmp_path / "output.ndjson"

        to_ndjson(sample_records, output_file)

        with open(output_file, encoding='utf-8-sig') as f:
            first_line = f.readline()

        # Should be compact (no indentation within the line)
        assert '  ' not in first_line


class TestXMLWriter:
    """Tests specific to XML writer."""

    def test_write_xml(self, tmp_path, sample_records):
        """Test basic XML writing."""
        output_file = tmp_path / "output.xml"

        to_xml(sample_records, output_file)

        assert output_file.exists()

        # Parse and verify
        from lxml import etree
        tree = etree.parse(str(output_file))
        root = tree.getroot()

        assert root.tag == 'data'
        records = root.findall('record')
        assert len(records) == 10

    def test_custom_element_names(self, tmp_path, sample_records):
        """Test custom root and record element names."""
        output_file = tmp_path / "output.xml"

        to_xml(sample_records, output_file, root_element='trainees', record_element='trainee')

        from lxml import etree
        tree = etree.parse(str(output_file))
        root = tree.getroot()

        assert root.tag == 'trainees'
        records = root.findall('trainee')
        assert len(records) == 10

    def test_xml_pretty_print(self, tmp_path, sample_records):
        """Test XML with pretty printing."""
        output_file = tmp_path / "output.xml"

        to_xml(sample_records, output_file, pretty=True)

        with open(output_file, encoding='utf-8-sig') as f:
            content = f.read()

        # Should have indentation
        assert '\n  ' in content or '\n    ' in content

    def test_xml_streaming(self, tmp_path, sample_records):
        """Test XML streaming mode."""
        output_file = tmp_path / "output.xml"

        to_xml(sample_records, output_file, stream=True)

        assert output_file.exists()

        # Verify it's valid XML
        from lxml import etree
        tree = etree.parse(str(output_file))
        root = tree.getroot()

        records = root.findall('record')
        assert len(records) == 10


class TestDatabaseWriter:
    """Tests specific to DatabaseWriter."""

    def test_database_writer_init(self, sample_records):
        """Test DatabaseWriter initialization."""
        from unittest.mock import Mock
        from dbtk.writers.database import DatabaseWriter

        # Create mock cursor with required attributes
        mock_cursor = Mock()
        mock_cursor.connection.driver.paramstyle = 'named'

        writer = DatabaseWriter(
            data=sample_records,
            target_cursor=mock_cursor,
            target_table='test_table'
        )

        assert writer.target_table == 'test_table'
        assert writer.batch_size == 1000
        assert writer.commit_frequency == 10000
        assert writer.paramstyle == 'named'

    def test_database_writer_custom_params(self, sample_records):
        """Test DatabaseWriter with custom batch and commit settings."""
        from unittest.mock import Mock
        from dbtk.writers.database import DatabaseWriter

        mock_cursor = Mock()
        mock_cursor.connection.driver.paramstyle = 'qmark'

        writer = DatabaseWriter(
            data=sample_records,
            target_cursor=mock_cursor,
            target_table='test_table',
            batch_size=500,
            commit_frequency=5000
        )

        assert writer.batch_size == 500
        assert writer.commit_frequency == 5000

    def test_database_writer_insert_statement_generation(self, sample_records):
        """Test that INSERT statement is generated correctly."""
        from unittest.mock import Mock
        from dbtk.writers.database import DatabaseWriter

        mock_cursor = Mock()
        mock_cursor.connection.driver.paramstyle = 'named'

        writer = DatabaseWriter(
            data=sample_records,
            target_cursor=mock_cursor,
            target_table='test_table'
        )

        # Verify INSERT statement was created
        assert writer.insert_sql is not None
        assert 'INSERT INTO test_table' in writer.insert_sql
        assert writer.param_names is not None
        assert len(writer.param_names) > 0

    def test_cursor_to_cursor_function(self, sample_records):
        """Test cursor_to_cursor convenience function."""
        from unittest.mock import Mock, patch
        from dbtk.writers.database import cursor_to_cursor

        mock_cursor = Mock()
        mock_cursor.connection.driver.paramstyle = 'format'

        # Patch DatabaseWriter.write to avoid actual execution
        with patch('dbtk.writers.database.DatabaseWriter.write', return_value=10):
            count = cursor_to_cursor(
                source_data=sample_records,
                target_cursor=mock_cursor,
                target_table='test_table'
            )

        assert count == 10

    def test_database_writer_different_paramstyles(self, sample_dicts):
        """Test DatabaseWriter handles different paramstyles."""
        from unittest.mock import Mock
        from dbtk.writers.database import DatabaseWriter

        paramstyles = ['qmark', 'format', 'named', 'pyformat', 'numeric']

        for paramstyle in paramstyles:
            mock_cursor = Mock()
            mock_cursor.connection.driver.paramstyle = paramstyle

            writer = DatabaseWriter(
                data=sample_dicts,
                target_cursor=mock_cursor,
                target_table='test_table'
            )

            # Should create writer without errors
            assert writer.paramstyle == paramstyle
            assert writer.insert_sql is not None

    def test_custom_element_names(self, tmp_path, sample_records):
        """Test custom root and record element names."""
        output_file = tmp_path / "output.xml"

        to_xml(sample_records, output_file, root_element='trainees', record_element='trainee')

        from lxml import etree
        tree = etree.parse(str(output_file))
        root = tree.getroot()

        assert root.tag == 'trainees'
        records = root.findall('trainee')
        assert len(records) == 10

    def test_xml_pretty_print(self, tmp_path, sample_records):
        """Test XML with pretty printing."""
        output_file = tmp_path / "output.xml"

        to_xml(sample_records, output_file, pretty=True)

        with open(output_file, encoding='utf-8-sig') as f:
            content = f.read()

        # Should have indentation
        assert '\n  ' in content or '\n    ' in content

    def test_xml_streaming(self, tmp_path, sample_records):
        """Test XML streaming mode."""
        output_file = tmp_path / "output.xml"

        to_xml(sample_records, output_file, stream=True)

        assert output_file.exists()

        # Verify it's valid XML
        from lxml import etree
        tree = etree.parse(str(output_file))
        root = tree.getroot()

        records = root.findall('record')
        assert len(records) == 10


class TestDatabaseWriter:
    """Tests specific to DatabaseWriter."""

    def test_database_writer_init(self, sample_records):
        """Test DatabaseWriter initialization."""
        from unittest.mock import Mock
        from dbtk.writers.database import DatabaseWriter

        # Create mock cursor with required attributes
        mock_cursor = Mock()
        mock_cursor.connection.driver.paramstyle = 'named'

        writer = DatabaseWriter(
            data=sample_records,
            target_cursor=mock_cursor,
            target_table='test_table'
        )

        assert writer.target_table == 'test_table'
        assert writer.batch_size == 1000
        assert writer.commit_frequency == 10000
        assert writer.paramstyle == 'named'

    def test_database_writer_custom_params(self, sample_records):
        """Test DatabaseWriter with custom batch and commit settings."""
        from unittest.mock import Mock
        from dbtk.writers.database import DatabaseWriter

        mock_cursor = Mock()
        mock_cursor.connection.driver.paramstyle = 'qmark'

        writer = DatabaseWriter(
            data=sample_records,
            target_cursor=mock_cursor,
            target_table='test_table',
            batch_size=500,
            commit_frequency=5000
        )

        assert writer.batch_size == 500
        assert writer.commit_frequency == 5000

    def test_database_writer_insert_statement_generation(self, sample_records):
        """Test that INSERT statement is generated correctly."""
        from unittest.mock import Mock
        from dbtk.writers.database import DatabaseWriter

        mock_cursor = Mock()
        mock_cursor.connection.driver.paramstyle = 'named'

        writer = DatabaseWriter(
            data=sample_records,
            target_cursor=mock_cursor,
            target_table='test_table'
        )

        # Verify INSERT statement was created
        assert writer.insert_sql is not None
        assert 'INSERT INTO test_table' in writer.insert_sql
        assert writer.param_names is not None
        assert len(writer.param_names) > 0

    def test_cursor_to_cursor_function(self, sample_records):
        """Test cursor_to_cursor convenience function."""
        from unittest.mock import Mock, patch
        from dbtk.writers.database import cursor_to_cursor

        mock_cursor = Mock()
        mock_cursor.connection.driver.paramstyle = 'format'

        # Patch DatabaseWriter.write to avoid actual execution
        with patch('dbtk.writers.database.DatabaseWriter.write', return_value=10):
            count = cursor_to_cursor(
                source_data=sample_records,
                target_cursor=mock_cursor,
                target_table='test_table'
            )

        assert count == 10

    def test_database_writer_different_paramstyles(self, sample_dicts):
        """Test DatabaseWriter handles different paramstyles."""
        from unittest.mock import Mock
        from dbtk.writers.database import DatabaseWriter

        paramstyles = ['qmark', 'format', 'named', 'pyformat', 'numeric']

        for paramstyle in paramstyles:
            mock_cursor = Mock()
            mock_cursor.connection.driver.paramstyle = paramstyle

            writer = DatabaseWriter(
                data=sample_dicts,
                target_cursor=mock_cursor,
                target_table='test_table'
            )

            # Should create writer without errors
            assert writer.paramstyle == paramstyle
            assert writer.insert_sql is not None

COMPRESSION_RECORDS = [
    {'id': 1, 'name': 'Alice', 'city': 'Springfield'},
    {'id': 2, 'name': 'Bob', 'city': 'Shelbyville'},
    {'id': 3, 'name': 'Carol', 'city': 'Capital City'},
]


class TestCompression:
    """Tests for compression support in file writers."""

    def test_csv_gz_extension_infers_gzip(self, tmp_path):
        """Extension .csv.gz triggers gzip compression automatically."""
        output_file = tmp_path / "output.csv.gz"
        to_csv(COMPRESSION_RECORDS, output_file)

        assert output_file.exists()
        import gzip
        with gzip.open(output_file, 'rt', encoding='utf-8') as f:
            first_line = f.readline()
        assert 'id' in first_line

    def test_csv_gz_roundtrip(self, tmp_path):
        """Write and read back a gzip-compressed CSV."""
        output_file = tmp_path / "output.csv.gz"
        to_csv(COMPRESSION_RECORDS, output_file)

        import gzip
        with gzip.open(output_file, 'rt', encoding='utf-8') as f:
            lines = f.readlines()
        # 1 header + 3 data rows
        assert len(lines) == 4

    def test_csv_explicit_gzip_no_extension(self, tmp_path):
        """Explicit compression='gzip' overrides extension inference."""
        output_file = tmp_path / "output.csv"
        to_csv(COMPRESSION_RECORDS, output_file, compression='gzip')

        import gzip
        with gzip.open(output_file, 'rt', encoding='utf-8') as f:
            first_line = f.readline()
        assert 'id' in first_line

    def test_csv_bz2_extension(self, tmp_path):
        """Extension .csv.bz2 triggers bz2 compression."""
        output_file = tmp_path / "output.csv.bz2"
        to_csv(COMPRESSION_RECORDS, output_file)

        import bz2
        with bz2.open(output_file, 'rt', encoding='utf-8') as f:
            first_line = f.readline()
        assert 'id' in first_line

    def test_csv_xz_extension(self, tmp_path):
        """Extension .csv.xz triggers lzma compression."""
        output_file = tmp_path / "output.csv.xz"
        to_csv(COMPRESSION_RECORDS, output_file)

        import lzma
        with lzma.open(output_file, 'rt', encoding='utf-8') as f:
            first_line = f.readline()
        assert 'id' in first_line

    def test_csv_no_compression_plain_extension(self, tmp_path):
        """Plain .csv extension writes uncompressed."""
        output_file = tmp_path / "output.csv"
        to_csv(COMPRESSION_RECORDS, output_file)

        with open(output_file, encoding='utf-8') as f:
            first_line = f.readline()
        assert 'id' in first_line

    def test_csv_compression_none_disables_inference(self, tmp_path):
        """compression=None writes plain text even with .gz extension."""
        output_file = tmp_path / "output.csv.gz"
        to_csv(COMPRESSION_RECORDS, output_file, compression=None)

        with open(output_file, encoding='utf-8') as f:
            first_line = f.readline()
        assert 'id' in first_line

    def test_json_gz_extension(self, tmp_path):
        """Extension .json.gz triggers gzip compression."""
        output_file = tmp_path / "output.json.gz"
        to_json(COMPRESSION_RECORDS, output_file)

        import gzip, json
        with gzip.open(output_file, 'rt', encoding='utf-8') as f:
            data = json.load(f)
        assert len(data) == 3
        assert data[0]['id'] == 1

    def test_ndjson_gz_extension(self, tmp_path):
        """Extension .ndjson.gz triggers gzip compression."""
        output_file = tmp_path / "output.ndjson.gz"
        to_ndjson(COMPRESSION_RECORDS, output_file)

        import gzip, json
        with gzip.open(output_file, 'rt', encoding='utf-8') as f:
            lines = f.readlines()
        assert len(lines) == 3
        assert json.loads(lines[0])['id'] == 1

    def test_csv_batch_writer_gz(self, tmp_path):
        """CSVWriter.write_batch() works with gzip compression."""
        output_file = tmp_path / "output.csv.gz"

        with CSVWriter(data=None, file=output_file) as writer:
            writer.write_batch(COMPRESSION_RECORDS[:2])
            writer.write_batch(COMPRESSION_RECORDS[2:])

        import gzip
        with gzip.open(output_file, 'rt', encoding='utf-8') as f:
            lines = f.readlines()
        assert len(lines) == 4  # 1 header + 3 data
